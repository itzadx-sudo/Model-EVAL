"""Parser tests — structure, omission detection, skip-section filtering, and
graceful handling of malformed / missing-column files."""

from __future__ import annotations

import openpyxl

from backend.hecvat_parser import parse_hecvat, skipped_refs


def test_parses_items(sample_hecvat, profile_path):
    items = parse_hecvat(sample_hecvat, profile_path)
    assert items, "expected at least some parsed items"
    refs = {it.ref for it in items}
    assert "DOCU-01" in refs
    # Section name tracked from the divider row.
    docu = next(it for it in items if it.ref == "DOCU-01")
    assert docu.section == "Documentation"
    assert docu.vendor_answer is not None


def test_blank_answer_is_none(sample_hecvat, profile_path):
    items = parse_hecvat(sample_hecvat, profile_path)
    docu03 = next(it for it in items if it.ref == "DOCU-03")
    assert docu03.vendor_answer is None  # -> omission downstream


def test_skip_sections_excluded_by_default(sample_hecvat, profile_path):
    items = parse_hecvat(sample_hecvat, profile_path)
    assert all(it.prefix != "HIPA" for it in items)
    # …but they exist in the file and are reported by skipped_refs.
    skipped = skipped_refs(sample_hecvat, profile_path)
    assert "HIPA-01" in skipped


def test_include_skipped_keeps_them(sample_hecvat, profile_path):
    items = parse_hecvat(sample_hecvat, profile_path, include_skipped=True)
    assert any(it.prefix == "HIPA" for it in items)


def test_missing_columns_does_not_crash(tmp_path, profile_path):
    """A sheet with fewer columns than the profile expects must not raise."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Organization"
    ws.cell(row=1, column=1, value=" Documentation")  # section header, 1 col only
    ws.cell(row=2, column=1, value="DOCU-01")  # id but no question/answer columns
    ws.cell(row=3, column=1, value="DOCU-02")
    ws.cell(row=3, column=2, value="A question with no answer column")
    path = tmp_path / "broken.xlsx"
    wb.save(path)

    items = parse_hecvat(str(path), profile_path)
    # DOCU-01 has no question text -> skipped; DOCU-02 parses with answer=None.
    refs = {it.ref for it in items}
    assert "DOCU-02" in refs
    assert next(it for it in items if it.ref == "DOCU-02").vendor_answer is None


def test_real_template_parses(profile_path):
    """The real EDUCAUSE template (structure only) parses without error."""
    import pathlib

    real = pathlib.Path(__file__).resolve().parent.parent / "Copy of HECVAT413.xlsx"
    if not real.exists():
        return  # optional — only runs if the template is present
    items = parse_hecvat(str(real), profile_path)
    # Organization sheet has GNRL + DOCU + THRD + ... questions.
    assert any(it.ref.startswith("DOCU-") for it in items)
