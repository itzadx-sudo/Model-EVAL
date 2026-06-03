"""Generate samples/sample_hecvat_template.xlsx — a synthetic, HECVAT-shaped
workbook with fabricated vendor answers so the benchmark actually exercises the
LLM end-to-end with no confidential data.

The real EDUCAUSE template ships with *blank* answer cells (structure only), so a
benchmark against it would be all omissions and never call the model. This sample
mirrors the real column layout (id | question | answer | additional info) and the
"section header rows start col A with a space" convention, then fills a handful of
Organization answers (some intentionally blank to exercise omission handling).

Run:  python samples/build_sample.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

# (ref, question, vendor_answer) — vendor_answer "" means blank -> omission.
ROWS = [
    ("__section__", "Documentation", None),
    (
        "DOCU-01",
        "Do you have a well-documented business continuity plan (BCP)?",
        "Yes. Our BCP is reviewed annually and was last tested in March 2026.",
    ),
    (
        "DOCU-02",
        "Do you have a well-documented disaster recovery (DR) plan?",
        "We maintain a DR plan with an RTO of 4 hours and RPO of 1 hour.",
    ),
    (
        "DOCU-03",
        "Have you undergone a SSAE 18/SOC 2 audit in the last 12 months?",
        "",  # blank -> omission, no LLM call
    ),
    (
        "DOCU-04",
        "Do you conform with a specific industry security standard (e.g. ISO 27001)?",
        "Yes, we are ISO 27001:2022 certified across all production environments.",
    ),
    (
        "DOCU-05",
        "Can you provide overall system and network architecture diagrams?",
        "Diagrams are available under NDA on request.",
    ),
    ("__section__", "Assessment of Third Parties", None),
    (
        "THRD-01",
        "Do you perform security assessments of third-party companies with access to data?",
        "Third parties are risk-assessed at onboarding but not re-assessed on a schedule.",
    ),
    (
        "THRD-02",
        "Do you require third parties to comply with your security requirements?",
        "Yes, security requirements are contractually mandated for all subprocessors.",
    ),
    ("__section__", "Application Security", None),
    (
        "APPL-01",
        "Does your application enforce password complexity and rotation requirements?",
        "We enforce 12-character minimums and MFA; rotation is not forced.",
    ),
    (
        "APPL-02",
        "Is data encrypted in transit using TLS 1.2 or higher?",
        "All traffic uses TLS 1.3; TLS 1.2 is the minimum accepted.",
    ),
    (
        "APPL-03",
        "Do you perform regular penetration testing of the application?",
        "",  # blank -> omission
    ),
    ("__section__", "Incident Response", None),
    (
        "HFIH-01",
        "Do you have a documented incident response plan and notify customers of breaches?",
        "We notify affected customers within 72 hours of confirming a breach.",
    ),
    (
        "HFIH-02",
        "Do you maintain audit logs sufficient to investigate a security incident?",
        "Logs are retained for 90 days and are immutable.",
    ),
    # A skip-section item to prove skip logic produces zero LLM calls.
    ("__section__", "HIPAA", None),
    (
        "HIPA-01",
        "Are you a HIPAA business associate?",
        "Not applicable — US-specific.",
    ),
]


def build(out_path: str = "samples/sample_hecvat_template.xlsx") -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Organization"

    # Mimic the real template's pre-amble rows so header detection is realistic.
    ws["A1"] = "HECVAT Solution Provider Response (synthetic sample — no confidential data)"
    ws["F2"] = "Version 4.1.3"

    row = 12
    for ref, question, answer in ROWS:
        if ref == "__section__":
            # Section divider: col A starts with a space, "Answer" label in col C.
            ws.cell(row=row, column=1, value=f" {question}")
            ws.cell(row=row, column=3, value="Answer")
            ws.cell(row=row, column=4, value="Additional Information")
        else:
            ws.cell(row=row, column=1, value=ref)
            ws.cell(row=row, column=2, value=question)
            if answer:
                ws.cell(row=row, column=3, value=answer)
        row += 1

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
