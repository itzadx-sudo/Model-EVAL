"""Parse a HECVAT .xlsx (Full or Lite) into a list of HecvatItem.

Column mappings and sheet names come from ``config/hecvat_profile.yaml`` — never
hardcoded. The parser walks each configured sheet row by row:

* Rows whose first column starts with the ``section_headers.marker`` (a space)
  are section dividers — they set the *current section name* (e.g. "Documentation").
* Rows whose first column matches a question-ID pattern (e.g. ``DOCU-01``) are
  questions — column indices for ID / text / vendor answer come from the profile.
* Everything else (instructions, blank rows) is ignored.

A blank vendor answer means the item is an *omission* — the runner flags it and
makes no LLM call (D16). Sections whose ID prefix is in ``skip_sections`` are
dropped here so they never reach the model (Agent guidance: zero LLM calls).
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import openpyxl
import yaml

from backend.models import HecvatItem

log = logging.getLogger("aegis.hecvat_parser")

# A HECVAT question id looks like THREE-or-FOUR uppercase letters, a dash, digits.
_ID_RE = re.compile(r"^[A-Z]{3,4}-\d+[A-Za-z]?$")


def _cell(row: tuple, idx: int) -> object:
    """Safely fetch a column from a row tuple — missing column returns None
    instead of raising, so short/ragged rows never crash the parser."""
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def load_profile(profile_path: str) -> dict:
    """Read the HECVAT profile YAML (column mappings, sheets, skip list)."""
    with open(profile_path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def _clean(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_hecvat(
    hecvat_path: str,
    profile_path: str,
    *,
    sheets: list[str] | None = None,
    include_skipped: bool = False,
) -> list[HecvatItem]:
    """Parse the HECVAT workbook into HecvatItem objects.

    Args:
        hecvat_path: path to the .xlsx file.
        profile_path: path to hecvat_profile.yaml.
        sheets: override the sheets to read; defaults to the profile primary
            sheet. Pass the profile's ``all_sheets`` to read everything.
        include_skipped: if True, keep items whose prefix is in skip_sections
            (used by tests to assert the skip count). Normal runs drop them.
    """
    profile = load_profile(profile_path)
    id_col = int(profile.get("question_id_col", 0))
    text_col = int(profile.get("question_text_col", 1))
    answer_col = int(profile.get("vendor_answer_col", 2))
    marker = profile.get("section_headers", {}).get("marker", " ")
    skip_prefixes = {p.strip().upper() for p in profile.get("skip_sections", [])}

    if sheets is None:
        sheets = [profile.get("primary_sheet", "Organization")]

    try:
        wb = openpyxl.load_workbook(hecvat_path, read_only=True, data_only=True)
    except Exception as exc:  # corrupt / unreadable file -> empty, don't crash
        log.error("Could not open HECVAT workbook %s: %s", hecvat_path, exc)
        return []

    items: list[HecvatItem] = []
    available = set(wb.sheetnames)

    for sheet_name in sheets:
        if sheet_name not in available:
            log.warning("Sheet %r not found in %s — skipping", sheet_name, Path(hecvat_path).name)
            continue
        try:
            ws = wb[sheet_name]
            current_section = sheet_name
            for row in ws.iter_rows(values_only=True):
                try:
                    if not row:
                        continue
                    raw_a = _cell(row, id_col)
                    if raw_a is None:
                        continue
                    cell = str(raw_a)
                    stripped = cell.strip()

                    # Section divider: col A starts with the marker (a space)
                    # and is not itself a question id.
                    if stripped and cell.startswith(marker) and not _ID_RE.match(stripped):
                        current_section = stripped
                        continue

                    if not _ID_RE.match(stripped):
                        continue

                    ref = stripped
                    question = _clean(_cell(row, text_col))
                    vendor_answer = _clean(_cell(row, answer_col))
                    if question is None:
                        # An id with no question text (e.g. a missing column) is
                        # not a usable item — skip it rather than crash.
                        log.debug("Row %s has no question text — skipping", ref)
                        continue

                    prefix = ref.split("-")[0].upper()
                    if prefix in skip_prefixes and not include_skipped:
                        continue

                    items.append(
                        HecvatItem(
                            ref=ref,
                            section=current_section,
                            question=question,
                            vendor_answer=vendor_answer,
                            sheet_name=sheet_name,
                        )
                    )
                except Exception as exc:  # one bad row never kills the whole parse
                    log.warning("Skipping malformed row in %s: %s", sheet_name, exc)
                    continue
        except Exception as exc:  # one bad sheet never kills the whole file
            log.warning("Skipping unreadable sheet %r: %s", sheet_name, exc)
            continue

    wb.close()
    return items


def skipped_refs(hecvat_path: str, profile_path: str, sheets: list[str] | None = None) -> list[str]:
    """Return the refs that *would* be skipped — used to assert zero LLM calls."""
    profile = load_profile(profile_path)
    skip_prefixes = {p.strip().upper() for p in profile.get("skip_sections", [])}
    every = parse_hecvat(hecvat_path, profile_path, sheets=sheets, include_skipped=True)
    return [it.ref for it in every if it.prefix in skip_prefixes]


if __name__ == "__main__":  # tiny smoke check
    import sys

    path = sys.argv[1] if len(sys.argv) > 1 else "samples/sample_hecvat_template.xlsx"
    prof = sys.argv[2] if len(sys.argv) > 2 else "config/hecvat_profile.yaml"
    parsed = parse_hecvat(path, prof)
    print(f"Parsed {len(parsed)} items from {Path(path).name}")
    for it in parsed[:10]:
        ans = (it.vendor_answer or "(blank)")[:40]
        print(f"  {it.ref:10} [{it.section[:20]:20}] {ans}")
